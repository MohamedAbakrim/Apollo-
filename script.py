import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains

driver = webdriver.Chrome()

extension_path = 'extension_path.crx'

chrome_options = webdriver.ChromeOptions()
chrome_options.add_extension(extension_path)

# Close the new tab & Switch back to the original tab
driver = webdriver.Chrome(options=chrome_options)
time.sleep(3)
# driver.switch_to.window(driver.window_handles[1])
# time.sleep(1)
# driver.close()
driver.switch_to.window(driver.window_handles[0])

driver.get("https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin")

firstname_input = driver.find_element(By.ID, 'username')
firstname_input.send_keys('mohammedabakrim@gmail.com') # Replace it with your email address
time.sleep(2)
password_input = driver.find_element(By.ID, 'password')
password_input.send_keys('#AbakriM065841#') # Replace it with your password
time.sleep(3)
button1 = driver.find_element(By.CLASS_NAME, 'login__form_action_container')
button1.click()
time.sleep(50)


# Sign-in Apollo
# driver.get("https://app.apollo.io/#/login?redirectTo=https%3A%2F%2Fapp.apollo.io%2F%23%2F")
driver.get("https://login.microsoftonline.com/common/oauth2/v2.0/authorize?client_id=f94ccf82-3918-4567-8a93-da0e5c2a51f7&login_hint=&prompt=select_account&redirect_uri=https%3A%2F%2Fapp.apollo.io%2Fapi%2Fv1%2Femail_accounts%2Fms_auth_callback&response_type=code&scope=openid+profile+user.read+offline_access&state=eyJ1c2VyX2FjY2VwdGVkX3Rlcm1zIjoidHJ1ZSIsIm1vZGUiOiJzZWxmX3NlcnZlX3dlYnNpdGVfc2lnbnVwIiwidXRtX21lZGl1bSI6ImNocm9tZV9leHRlbnNpb25fc3RvcmUiLCJ1dG1fc291cmNlIjoiY2hyb21lX2V4dGVuc2lvbl9zdG9yZSIsImluaXRpYWxfdXRtX21lZGl1bSI6ImNocm9tZV9leHRlbnNpb25fc3RvcmUiLCJpbml0aWFsX3V0bV9zb3VyY2UiOiJjaHJvbWVfZXh0ZW5zaW9uX3N0b3JlIiwiaW5pdGlhbF9yZWZlcnJlciI6Imh0dHBzOi8vYXBwLmFwb2xsby5pby8iLCJwcmljaW5nX3ZhcmlhbnQiOiIyM1ExX0VDX1k0OSIsInJlZGlyZWN0X3RvIjoiaHR0cHM6Ly93d3cuYXBvbGxvLmlvL3NpZ251cC1zdWNjZXNzLXdpdGgtbWljcm9zb2Z0LXN0YW5kYXJkPyIsInpwX3NvdXJjZSI6InNpZ251cCJ9&username=&sso_reload=true")
time.sleep(10)

# driver.get("https://www.apollo.io/sign-up?utm_source=chrome_extension_store&utm_medium=chrome_extension_store")
# button = driver.find_element(By.CLASS_NAME, 'PrivateSwitchBase-input')
# button.click()
# button2 = driver.find_elements(By.CLASS_NAME, 'MuiButton-outlinedSizeMedium')
# button2[2].click()
# time.sleep(3)

username_field = driver.find_element(By.ID, "i0116")
username_field.send_keys("drtdelicia@outlook.com")
time.sleep(2)

buttonNext = driver.find_element(By.ID, 'idSIButton9')
buttonNext.click()
time.sleep(15)

password_field = driver.find_element(By.ID, "i0118")
password_field.send_keys("o4hklmfU")
time.sleep(2)

password_field.send_keys(Keys.RETURN)
time.sleep(10)
buttonBack = driver.find_element(By.ID, 'idBtn_Back')
buttonBack.click()

time.sleep(5)


# Load the Excel file
excel_file_path = 'excel.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active


for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=4):
    link_cell, phone_cell, plus_phone_cell, email_cell = row[0], row[1], row[2], row[3]

    link = link_cell.value
    print(link)

    driver.get(link)
    time.sleep(30)

    try:
        buttonClick = driver.find_element(By.CLASS_NAME, 'x_LQDkG')
        buttonClick.click()
        time.sleep(5)
    except:
        pass    

    try:
        buttonClose = driver.find_element(By.CLASS_NAME, 'mdi-close')
        buttonClose.click()
        time.sleep(2)
    except:
        pass  

    try:
        view_mobile_number = driver.find_element(By.CLASS_NAME, 'x_LQDkG') 
        view_mobile_number.click()
        time.sleep(3)
    except:
        pass

    try:
        test = driver.find_elements(By.CLASS_NAME, 'x_pRbdE')
        for i in test:
            if i.text == 'Mobile':
                i.click()
            else:
                pass
        time.sleep(10)
    except:
        pass

    # try:
    #     RequestNumber = driver.find_elements(By.CLASS_NAME, 'x_OotKe')
    #     RequestNumber[0].click()
    #     time.sleep(2)
    # except:
    #     pass    

    try:
        email = driver.find_element(By.CLASS_NAME, 'x_GxQlI').text
    except:
        email = "None"

    try:
        phone = driver.find_elements(By.CLASS_NAME, 'x_XitPs')[0].text
    except:
        phone = "None"

    print(email, phone)

    phone_cell.value = phone 
    email_cell.value = email

    phones = []
    try:

        elements = driver.find_elements(By.CLASS_NAME, 'x_XitPs')
        for element in elements:
            text = element.text
            phones.append(text)

        print("##############11", phones)

    except IndexError:
            pass
    

    phones_without_first_element = phones[1:]
    phone_string = '\n'.join(map(str, phones_without_first_element))
    plus_phone_cell.value = phone_string


    workbook.save('excel.xlsx')


driver.quit()

